# Script para extraer estructura del Excel (hojas, columnas, primeras filas)
# Ejecutar: python leer_excel.py

import json
import sys

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

EXCEL_PATH = r"C:\Users\Lenovo\Proyectos\Recetas de Helados\FORMULADOR BASE  CURSO GRATIS YT ESCUELA DUBOVIK .xlsx"
OUT_PATH = r"C:\Users\Lenovo\Proyectos\Recetas de Helados\estructura_excel.txt"

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    lines = []
    lines.append("=== ESTRUCTURA DEL EXCEL ===\n")
    lines.append("Hojas: " + ", ".join(wb.sheetnames) + "\n")
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append("\n--- Hoja: " + name + " ---")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("(vacía)")
            continue
        # Primera fila = cabeceras
        headers = [str(c) if c is not None else "" for c in rows[0]]
        lines.append("Columnas: " + " | ".join(headers))
        lines.append("Filas de datos (primeras 15):")
        for i, row in enumerate(rows[1:16]):
            vals = [str(v)[:50] if v is not None else "" for v in row]
            lines.append("  " + str(i+1) + ": " + " | ".join(vals))
        if len(rows) > 16:
            lines.append("  ... (" + str(len(rows)-1) + " filas más)")
    wb.close()
    text = "\n".join(lines)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write(text)
    print("OK. Estructura guardada en:", OUT_PATH)

if __name__ == "__main__":
    main()
